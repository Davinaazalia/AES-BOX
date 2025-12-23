# app.py - Frontend + Backend Combined
import os
from flask import Flask, request, render_template, redirect, url_for, flash, send_from_directory, send_file, jsonify
from flask_cors import CORS
from werkzeug.utils import secure_filename
import pandas as pd
from datetime import datetime

from core.sbox_generator import generate_sbox, generate_sbox_from_matrix
from core.sbox_validator import is_bijective, is_balanced, check_sac, differential_uniformity, nonlinearity, calculate_sac_value
from core.utils import allowed_file, bit_balance, avalanche_test
from core.sbox_examples import SBOX1, SBOX2, SBOX3
from core.matrix_explorer import explore_affine_candidates, get_top_candidates
from services.excel_service import read_sbox_from_excel, read_sbox_from_file, export_sbox_to_excel, export_analysis_to_excel, GENERATED_DIR
from services.image_encrypt import apply_subbytes_to_image, image_entropy, npcr, histogram_counts, histogram_rgb, uaci
from services.aes_service import (
    decrypt_file,
    decrypt_text,
    encrypt_file,
    encrypt_text,
)
import json


BASE_DIR = os.path.dirname(__file__)
OUTPUT_ROOT = os.path.join(BASE_DIR, 'outputs')
UPLOAD_FOLDER = os.path.join(OUTPUT_ROOT, 'uploaded_sboxes')
GENERATED_FOLDER = os.path.join(OUTPUT_ROOT, 'generated_sboxes')
ENCRYPTED_FOLDER = os.path.join(OUTPUT_ROOT, 'encrypted_images')
AES_CIPHER_FOLDER = os.path.join(OUTPUT_ROOT, 'aes_cipher')
AES_PLAIN_FOLDER = os.path.join(OUTPUT_ROOT, 'aes_plain')
AES_VISUAL_FOLDER = os.path.join(OUTPUT_ROOT, 'aes_visual')
FRONTEND_ROOT = os.path.join(BASE_DIR, '..', 'frontend')
FRONTEND_BUILD = os.path.join(FRONTEND_ROOT, 'dist')
TEMPLATE_FOLDER = os.path.join(BASE_DIR, 'templates')

ALLOWED_EXCEL = {'xlsx', 'xls'}
ALLOWED_SBOX = {'xlsx', 'xls', 'csv', 'txt', 'json'}
ALLOWED_IMAGES = {'png', 'jpg', 'jpeg'}

for path in [UPLOAD_FOLDER, GENERATED_FOLDER, ENCRYPTED_FOLDER, AES_CIPHER_FOLDER, AES_PLAIN_FOLDER, AES_VISUAL_FOLDER]:
    os.makedirs(path, exist_ok=True)

app = Flask(__name__, template_folder=TEMPLATE_FOLDER)
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = 'ganti_dengan_secret_key_yang_aman'


def validate_sbox_format(flat):
    if len(flat) != 256:
        return False, "S-box harus 256 nilai (16x16)."
    for v in flat:
        if not (0 <= int(v) <= 255):
            return False, f"Nilai {v} di luar rentang 0..255."
    if len(set(flat)) != 256:
        dup = 256 - len(set(flat))
        return False, f"S-box tidak bijektif: terdapat {dup} duplikat."
    return True, "S-box valid dan bijektif."


@app.route('/')
def landing_page():
    return render_template('landing.html')


@app.route('/analyzer')
def index():
    return render_template('analyzer.html')


@app.route('/aes')
def aes_page():
    return render_template('aes.html')


@app.route('/batch-analyzer')
def batch_analyzer():
    return render_template('batch_sbox_analyzer.html')


@app.route('/outputs')
def view_outputs():
    """Display list of generated outputs"""
    outputs = {
        'generated_sboxes': [],
        'encrypted_images': [],
        'aes_cipher': [],
        'aes_plain': [],
        'aes_visual': []
    }
    
    # List files in each output directory
    for output_type, folder_path in [
        ('generated_sboxes', GENERATED_FOLDER),
        ('encrypted_images', ENCRYPTED_FOLDER),
        ('aes_cipher', AES_CIPHER_FOLDER),
        ('aes_plain', AES_PLAIN_FOLDER),
        ('aes_visual', AES_VISUAL_FOLDER)
    ]:
        try:
            if os.path.exists(folder_path):
                files = os.listdir(folder_path)
                outputs[output_type] = sorted(files, reverse=True)[:20]  # Latest 20 files
        except:
            pass
    
    return render_template('outputs.html', outputs=outputs)


# Serve built React app (production) at /ui/*
@app.route('/ui', defaults={'path': ''})
@app.route('/ui/<path:path>')
def serve_react(path):
    # If dist exists, serve static files; otherwise fallback to dev server
    if os.path.isdir(FRONTEND_BUILD):
        candidate = os.path.join(FRONTEND_BUILD, path)
        if path and os.path.exists(candidate) and os.path.isfile(candidate):
            return send_from_directory(FRONTEND_BUILD, path)
        index_path = os.path.join(FRONTEND_BUILD, 'index.html')
        if os.path.exists(index_path):
            return send_file(index_path)
    # fallback to dev server
    return redirect(f'http://localhost:5173/{path}')


@app.route('/analyze', methods=['POST'])
def analyze():
    report = {}

    if 'sbox' not in request.files:
        flash('File S-box tidak ditemukan.')
        return redirect(url_for('index'))

    sbox_file = request.files['sbox']
    if sbox_file.filename == '':
        flash('Tidak ada file S-box yang dipilih.')
        return redirect(url_for('index'))

    if not allowed_file(sbox_file.filename, ALLOWED_SBOX):
        flash('Format file harus .xlsx, .xls, .csv, .txt, atau .json')
        return redirect(url_for('index'))

    sbox_fname = secure_filename(sbox_file.filename)
    sbox_path = os.path.join(UPLOAD_FOLDER, sbox_fname)
    sbox_file.save(sbox_path)

    try:
        flat, mat = read_sbox_from_file(sbox_path)
    except Exception as e:
        flash(f"Error membaca file: {e}")
        return redirect(url_for('index'))

    ok, msg = validate_sbox_format(flat)
    report['valid'] = ok
    report['message'] = msg
    report['matrix_html'] = pd.DataFrame(mat).to_html(classes='table table-sm', header=False, index=False)

    report['bit_balance'] = bit_balance(flat)
    report['avalanche'] = {f'bit_{b}': round(avalanche_test(flat, flip_bit=b), 4) for b in range(8)}

    report['bijective'] = is_bijective(flat)
    report['balanced'] = is_balanced(flat)
    report['sac'] = check_sac(flat)
    report['differential_uniformity'] = differential_uniformity(flat)
    report['nonlinearity'] = nonlinearity(flat)

    if 'sample_img' in request.files and request.files['sample_img'].filename != '':
        imgf = request.files['sample_img']
        if imgf and allowed_file(imgf.filename, ALLOWED_IMAGES):
            img_name = secure_filename(imgf.filename)
            img_path = os.path.join(UPLOAD_FOLDER, img_name)
            imgf.save(img_path)

            cipher_name = 'cipher_' + img_name
            cipher_path = os.path.join(ENCRYPTED_FOLDER, cipher_name)
            apply_subbytes_to_image(img_path, flat, cipher_path)

            # Create modified plaintext for NPCR
            mod_plain = os.path.join(ENCRYPTED_FOLDER, 'plain_mod_' + img_name)
            cipher_mod = os.path.join(ENCRYPTED_FOLDER, 'cipher_mod_' + img_name)

            from PIL import Image
            import numpy as np

            im = Image.open(img_path).convert('RGB')
            arr = np.array(im)
            arr2 = arr.copy()
            arr2[0, 0, 0] = (int(arr2[0, 0, 0]) + 1) % 256
            Image.fromarray(arr2).save(mod_plain)
            apply_subbytes_to_image(mod_plain, flat, cipher_mod)

            report['image_cipher'] = cipher_name
            report['entropy'] = round(image_entropy(cipher_path), 6)
            report['npcr'] = round(npcr(cipher_path, cipher_mod), 6)
            # Histogram (grayscale) for plaintext and cipher
            report['hist_plain'] = histogram_counts(img_path)
            report['hist_cipher'] = histogram_counts(cipher_path)
            # RGB Histogram for plaintext and cipher
            report['hist_rgb_plain'] = histogram_rgb(img_path)
            report['hist_rgb_cipher'] = histogram_rgb(cipher_path)

    return render_template('index.html', report=report)


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.route('/encrypted/<filename>')
def encrypted_file(filename):
    return send_from_directory(ENCRYPTED_FOLDER, filename)


# Serve generated analysis/excel files
@app.route('/generated/<filename>')
def generated_file(filename):
    return send_from_directory(GENERATED_FOLDER, filename)


@app.route('/aes-cipher/<path:filename>')
def aes_cipher_file(filename):
    return send_from_directory(AES_CIPHER_FOLDER, filename)


@app.route('/aes-plain/<path:filename>')
def aes_plain_file(filename):
    return send_from_directory(AES_PLAIN_FOLDER, filename)

@app.route('/aes-visual/<path:filename>')
def aes_visual_file(filename):
    return send_from_directory(AES_VISUAL_FOLDER, filename)


@app.route('/download-example/<which>')
def download_example(which):
    which = which.lower()
    if which == 'sbox1':
        data = SBOX1
    elif which == 'sbox2':
        data = SBOX2
    elif which == 'sbox3':
        data = SBOX3
    else:
        return "Unknown example", 404

    flat = [v for row in data for v in row]
    fname = f"{which}.xlsx"
    path = export_sbox_to_excel(flat, filename=fname)
    return send_file(path, as_attachment=True, download_name=fname)


@app.route('/generate-example-sbox')
def generate_example_sbox():
    sbox = generate_sbox()
    return jsonify({"sbox": sbox})


@app.route('/api/', methods=['GET'])
def api_home():
    return jsonify({"message": "Backend S-Box API berjalan!"})


@app.route('/api/generate-sbox', methods=['GET'])
def api_generate():
    # Support random generation with seed parameter
    use_random = request.args.get('random', default='false').lower() == 'true'
    seed = request.args.get('seed', default=None, type=int)
    
    if use_random:
        from core.matrix_explorer import explore_affine_candidates
        if seed is not None:
            import numpy as np
            np.random.seed(seed)
        results = explore_affine_candidates(n_candidates=1, seed=seed)
        sbox = results[0]['sbox']
        return jsonify({
            "sbox": sbox,
            "random": True,
            "seed": seed,
            "metrics": {
                "bijective": results[0]['bijective'],
                "balanced": results[0]['balanced'],
                "sac": results[0]['sac'],
                "nonlinearity": results[0]['nonlinearity'],
                "diff_uniformity": results[0]['diff_uniformity']
            }
        })
    else:
        sbox = generate_sbox()
        return jsonify({"sbox": sbox, "random": False})


@app.route('/api/generate-sbox-excel', methods=['GET'])
def api_generate_excel():
    use_random = request.args.get('random', default='false').lower() == 'true'
    seed = request.args.get('seed', default=None, type=int)
    
    if use_random:
        from core.matrix_explorer import explore_affine_candidates
        results = explore_affine_candidates(n_candidates=1, seed=seed)
        sbox = results[0]['sbox']
    else:
        sbox = generate_sbox()
    
    path = export_sbox_to_excel(sbox)
    fname = os.path.basename(path)
    return send_file(path, as_attachment=True, download_name=fname)


@app.route('/api/validate-sbox', methods=['POST'])
def api_validate_sbox():
    data = request.get_json()
    sbox = data.get("sbox")
    result = {
        "bijective": is_bijective(sbox),
        "balanced": is_balanced(sbox),
        "sac": check_sac(sbox),
        "differential_uniformity": differential_uniformity(sbox),
        "nonlinearity": nonlinearity(sbox)
    }
    return jsonify(result)


@app.route('/api/validate-sbox-debug', methods=['GET'])
def api_validate_sbox_debug():
    sbox = generate_sbox()
    result = {
        "bijective": is_bijective(sbox),
        "balanced": is_balanced(sbox),
        "sac": check_sac(sbox),
        "differential_uniformity": differential_uniformity(sbox),
        "nonlinearity": nonlinearity(sbox)
    }
    return jsonify(result)


@app.route('/api/analyze', methods=['POST'])
def api_analyze_sbox():
    """API endpoint for S-Box analysis - returns JSON"""
    try:
        if 'sbox' not in request.files:
            return jsonify({'error': 'File S-box tidak ditemukan'}), 400

        sbox_file = request.files['sbox']
        if sbox_file.filename == '':
            return jsonify({'error': 'Tidak ada file S-box yang dipilih'}), 400

        if not allowed_file(sbox_file.filename, ALLOWED_SBOX):
            return jsonify({'error': 'Format file harus .xlsx, .xls, .csv, .txt, atau .json'}), 400

        sbox_fname = secure_filename(sbox_file.filename)
        sbox_path = os.path.join(UPLOAD_FOLDER, sbox_fname)
        sbox_file.save(sbox_path)
        
        print(f'📁 Reading S-box from: {sbox_path}')
        try:
            flat, mat = read_sbox_from_file(sbox_path)
        except ValueError as ve:
            return jsonify({
                'error': str(ve),
                'valid': False
            }), 400
        except Exception as ex:
            return jsonify({
                'error': f'Failed to read S-box file: {str(ex)}. Try exporting as CSV or JSON instead.',
                'valid': False
            }), 400
        print(f'✓ Read S-box: {len(flat)} values')

        ok, msg = validate_sbox_format(flat)
        bit_bal = bit_balance(flat)  # Returns list of 8 values
        report = {
            'valid': ok,
            'message': msg,
            'matrix': mat.tolist(),
            'sbox': [int(x) for x in flat],
            'bit_balance': float(sum(bit_bal)) / 8.0,  # Average bit balance
            'bit_balance_per_bit': [int(x) for x in bit_bal],  # Per-bit details
            'avalanche': {f'bit_{b}': round(avalanche_test(flat, flip_bit=b), 4) for b in range(8)},
            'bijective': bool(is_bijective(flat)),
            'balanced': bool(is_balanced(flat)),
            'sac': bool(check_sac(flat)),
            'differential_uniformity': int(differential_uniformity(flat)),
            'nonlinearity': int(nonlinearity(flat))
        }

        # Optional image analysis - ALWAYS run even if S-box invalid
        if 'sample_img' in request.files and request.files['sample_img'].filename != '':
            imgf = request.files['sample_img']
            if imgf and allowed_file(imgf.filename, ALLOWED_IMAGES):
                try:
                    img_name = secure_filename(imgf.filename)
                    img_path = os.path.join(UPLOAD_FOLDER, img_name)
                    imgf.save(img_path)

                    cipher_name = 'cipher_' + img_name
                    cipher_path = os.path.join(ENCRYPTED_FOLDER, cipher_name)
                    apply_subbytes_to_image(img_path, flat, cipher_path)

                    # Create modified plaintext for NPCR
                    from PIL import Image
                    import numpy as np
                    im = Image.open(img_path).convert('RGB')
                    arr = np.array(im)
                    arr2 = arr.copy()
                    arr2[0, 0, 0] = (int(arr2[0, 0, 0]) + 1) % 256
                    mod_plain = os.path.join(ENCRYPTED_FOLDER, 'plain_mod_' + img_name)
                    Image.fromarray(arr2).save(mod_plain)
                    
                    cipher_mod = os.path.join(ENCRYPTED_FOLDER, 'cipher_mod_' + img_name)
                    apply_subbytes_to_image(mod_plain, flat, cipher_mod)

                    # Get histogram data and ensure JSON-safe
                    hist_rgb_plain_data = histogram_rgb(img_path)
                    hist_rgb_cipher_data = histogram_rgb(cipher_path)
                    hist_plain_data = histogram_counts(img_path)
                    hist_cipher_data = histogram_counts(cipher_path)

                    # Convert to Python lists if numpy
                    report['image_analysis'] = {
                        'image_name': img_name,
                        'cipher_name': cipher_name,
                        'entropy': float(round(image_entropy(cipher_path), 6)),
                        'npcr': float(round(npcr(cipher_path, cipher_mod), 6)),
                        'hist_plain': [int(x) for x in hist_plain_data],
                        'hist_cipher': [int(x) for x in hist_cipher_data],
                        'hist_rgb_plain': {
                            'r': [int(x) for x in hist_rgb_plain_data.get('r', [])],
                            'g': [int(x) for x in hist_rgb_plain_data.get('g', [])],
                            'b': [int(x) for x in hist_rgb_plain_data.get('b', [])]
                        },
                        'hist_rgb_cipher': {
                            'r': [int(x) for x in hist_rgb_cipher_data.get('r', [])],
                            'g': [int(x) for x in hist_rgb_cipher_data.get('g', [])],
                            'b': [int(x) for x in hist_rgb_cipher_data.get('b', [])]
                        }
                    }
                except Exception as e:
                    report['image_error'] = str(e)
                    import traceback
                    traceback.print_exc()

        return jsonify(report)
    
    except Exception as e:
        print(f'❌ CRITICAL ERROR in /api/analyze: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Unexpected error: {str(e)}',
            'valid': False
        }), 500


@app.route('/api/sbox/generate-from-matrix', methods=['POST'])
def api_generate_sbox_from_matrix():
    """Generate S-box from custom matrix and constant"""
    try:
        data = request.get_json()
        matrix = data.get('matrix')
        constant = data.get('constant', '63')
        
        if not matrix or len(matrix) != 8:
            return jsonify({'error': 'Matrix harus 8x8'}), 400
        
        for row in matrix:
            if len(row) != 8:
                return jsonify({'error': 'Setiap row harus 8 bit'}), 400
        
        sbox = generate_sbox_from_matrix(matrix, constant)
        
        # Compute cryptographic metrics
        sac_value = calculate_sac_value(sbox)
        metrics = {
            'nl': nonlinearity(sbox),
            'sac': round(sac_value, 4),
            'bic_sac': round(sac_value, 4),  # Simplified - can add BIC-SAC later
            'lap': 16,  # Placeholder - needs LAP implementation
            'dap_prob': round(differential_uniformity(sbox) / 256, 4),  # Approximate DAP
            'du': differential_uniformity(sbox),
            'bic_nl': nonlinearity(sbox),  # Simplified - same as NL for now
            'alg_deg': 7,  # Placeholder - needs algebraic degree implementation
            'tg': 7.9731,  # Placeholder - needs transparency order implementation
            'bijective': is_bijective(sbox),
            'balanced': is_balanced(sbox),
            'sac_pass': check_sac(sbox)
        }
        
        return jsonify({
            'sbox': sbox,
            'matrix': matrix,
            'constant': constant,
            'metrics': metrics
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sbox/export-excel', methods=['POST'])
def api_export_sbox_excel():
    """Export S-box and analysis to Excel"""
    try:
        data = request.get_json()
        sbox = data.get('sbox')
        matrix = data.get('matrix')
        constant = data.get('constant', '63')
        preset = data.get('preset', 'custom')
        
        if not sbox:
            return jsonify({'error': 'S-box required'}), 400
        
        # Prepare data for export
        export_data = {
            'sbox': sbox,
            'matrix': matrix,
            'metrics': {
                'preset': preset,
                'constant': f'0x{constant}',
                'bijective': is_bijective(sbox),
                'balanced': is_balanced(sbox),
                'sac': check_sac(sbox),
                'differential_uniformity': differential_uniformity(sbox),
                'nonlinearity': nonlinearity(sbox)
            }
        }
        
        filepath = export_analysis_to_excel(export_data)
        filename = os.path.basename(filepath)
        
        return jsonify({
            'filename': filename,
            'download_url': url_for('generated_file', filename=filename)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/batch-export-excel', methods=['POST'])
def api_batch_export_excel():
    """Export batch analysis results to Excel"""
    try:
        results = request.get_json() or []
        if not results:
            return jsonify({'error': 'No results to export'}), 400
        
        # Validate and clean data
        cleaned_results = []
        for result in results:
            if isinstance(result, dict):
                # Extract sbox (should be a list of 256 ints)
                sbox = result.get('sbox', [])
                if isinstance(sbox, list) and len(sbox) == 256:
                    # Convert to ints if they're not already
                    try:
                        sbox = [int(x) for x in sbox]
                    except (ValueError, TypeError):
                        continue
                    
                    # Extract matrix (should be list of lists)
                    matrix = result.get('matrix', [])
                    if not isinstance(matrix, list):
                        matrix = []
                    
                    cleaned_results.append({
                        'name': result.get('name', 'Unknown'),
                        'sbox': sbox,
                        'matrix': matrix,
                        'metrics': result.get('metrics', {})
                    })
        
        if not cleaned_results:
            return jsonify({'error': 'No valid S-box data found'}), 400
        
        # Create workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        
        # Header row
        ws_summary.append(['Matrix', 'Bijective', 'Balanced', 'NL', 'SAC', 'BIC-NL', 'BIC-SAC', 'LAP', 'DAP', 'DU', 'ALG-DEG', 'TG', 'Score'])
        
        # Style header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Helper function to calculate score
        def calc_score(metrics):
            try:
                if not isinstance(metrics, dict):
                    return 0
                nl = float(metrics.get('nl', 0) or 0) / 112.0
                sac = float(metrics.get('sac', 0) or 0) / 0.5
                bic_nl = float(metrics.get('bic_nl', 0) or 0) / 112.0
                du_val = float(metrics.get('du', 1) or 1)
                du = 1.0 / (1.0 + du_val)
                tg = float(metrics.get('tg', 0) or 0) / 0.5
                score = (nl * 0.3 + sac * 0.2 + bic_nl * 0.2 + du * 0.15 + tg * 0.15) * 100
                return max(0, min(100, score))  # Clamp to 0-100
            except:
                return 0
        
        # Add data rows
        for result in cleaned_results:
            metrics = result.get('metrics', {})
            score = calc_score(metrics)
            
            ws_summary.append([
                result.get('name', ''),
                'Yes' if metrics.get('bijective') else 'No',
                'Yes' if metrics.get('balanced') else 'No',
                int(metrics.get('nl', 0) or 0),
                round(float(metrics.get('sac', 0) or 0), 4),
                int(metrics.get('bic_nl', 0) or 0),
                round(float(metrics.get('bic_sac', 0) or 0), 4),
                int(metrics.get('lap', 0) or 0),
                round(float(metrics.get('dap_prob', 0) or 0), 4),
                int(metrics.get('du', 0) or 0),
                int(metrics.get('alg_deg', 0) or 0),
                round(float(metrics.get('tg', 0) or 0), 4),
                round(score, 2)
            ])
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws_summary.column_dimensions[col].width = 12
        
        # Sheet 2: S-boxes (16x16 matrices)
        ws_sboxes = wb.create_sheet('S-Boxes')
        row = 1
        for result in cleaned_results:
            sbox = result.get('sbox', [])
            if sbox:
                # Add label
                ws_sboxes[f'A{row}'] = result.get('name', '')
                ws_sboxes[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # Add 16x16 matrix
                for i in range(16):
                    for j in range(16):
                        idx = i * 16 + j
                        if idx < len(sbox):
                            cell = ws_sboxes.cell(row=row+i, column=j+1)
                            cell.value = int(sbox[idx])
                            cell.alignment = Alignment(horizontal='center')
                
                row += 17  # 16 rows + 1 blank row
        
        # Sheet 3: Matrices (affine transformation matrices)
        ws_matrices = wb.create_sheet('Matrices')
        row = 1
        for result in cleaned_results:
            matrix = result.get('matrix', [])
            if matrix and len(matrix) > 0:
                # Add label
                ws_matrices[f'A{row}'] = result.get('name', '')
                ws_matrices[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # Add matrix (usually 8x8)
                for i in range(len(matrix)):
                    if isinstance(matrix[i], list):
                        for j in range(len(matrix[i])):
                            cell = ws_matrices.cell(row=row+i, column=j+1)
                            try:
                                cell.value = int(matrix[i][j])
                            except (ValueError, TypeError):
                                cell.value = matrix[i][j]
                            cell.alignment = Alignment(horizontal='center')
                
                row += len(matrix) + 1
        
        # Save file
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'batch_analysis_{ts}.xlsx'
        filepath = os.path.join(GENERATED_FOLDER, filename)
        
        print(f'Saving Excel to: {filepath}')
        wb.save(filepath)
        print(f'Excel saved successfully')
        
        # Return file as response
        return send_file(filepath, 
                        as_attachment=True,
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f'Error in batch export: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to generate Excel: {str(e)}'}), 500


@app.route('/api/aes/text/encrypt', methods=['POST'])
def api_aes_encrypt_text():
    data = request.get_json() or {}
    plaintext = data.get('plaintext', '')
    key_hex = data.get('key')
    iv_hex = data.get('iv')
    if not key_hex:
        return jsonify({"error": "Key (hex) wajib diisi"}), 400
    try:
        cipher_b64, iv_used = encrypt_text(plaintext, key_hex, iv_hex)
        return jsonify({"ciphertext": cipher_b64, "iv": iv_used})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route('/api/aes/text/decrypt', methods=['POST'])
def api_aes_decrypt_text():
    data = request.get_json() or {}
    cipher_b64 = data.get('ciphertext', '').strip()
    key_hex = data.get('key', '').strip()
    iv_hex = data.get('iv', '').strip()
    
    if not key_hex or not cipher_b64 or not iv_hex:
        return jsonify({"error": "ciphertext, key, dan iv wajib diisi"}), 400
    
    try:
        # Debug: validate inputs
        import base64
        try:
            ct_bytes = base64.b64decode(cipher_b64)
            print(f'✓ Ciphertext valid base64, decoded length: {len(ct_bytes)} bytes')
            
            # Check if decoded correctly
            if len(ct_bytes) == 0:
                return jsonify({
                    "error": "Base64 ciphertext kosong atau tidak valid. "
                             "Pastikan Anda mengcopy seluruh ciphertext dari hasil encrypt."
                }), 400
            
            if len(ct_bytes) < 16:
                return jsonify({
                    "error": f"Base64 ciphertext terlalu pendek ({len(ct_bytes)} bytes decoded). "
                             f"Harusnya minimal 16 bytes. "
                             f"Kemungkinan: ciphertext tidak lengkap atau tidak valid. "
                             f"Pastikan ciphertext yang dicopy penuh."
                }), 400
            
            if len(ct_bytes) % 16 != 0:
                return jsonify({
                    "error": f"Base64 ciphertext decoded ke {len(ct_bytes)} bytes, "
                             f"yang bukan kelipatan 16. "
                             f"Kemungkinan: base64 string corrupt atau tidak valid."
                }), 400
                
        except Exception as e:
            return jsonify({
                "error": f"Ciphertext harus valid base64: {str(e)}. "
                         f"Pastikan tidak ada spasi atau karakter aneh."
            }), 400
        
        # Validate IV format
        try:
            iv_bytes = bytes.fromhex(iv_hex)
            if len(iv_bytes) != 16:
                return jsonify({"error": f"IV harus 16 bytes, diterima {len(iv_bytes)} bytes"}), 400
            print(f'✓ IV valid hex, length: {len(iv_bytes)} bytes')
        except ValueError as e:
            return jsonify({"error": f"IV harus hex format valid: {str(e)}"}), 400
        
        # Try to determine key type
        try:
            key_bytes = bytes.fromhex(key_hex)
            if len(key_bytes) in (16, 24, 32):
                print(f'✓ Key adalah hex, length: {len(key_bytes)} bytes')
            else:
                print(f'Key hex length invalid ({len(key_bytes)}), akan diperlakukan sebagai passphrase')
        except:
            print(f'Key bukan hex, akan diderivasi dari passphrase: {key_hex}')
        
        plaintext = decrypt_text(cipher_b64, key_hex, iv_hex)
        return jsonify({"plaintext": plaintext})
    
    except ValueError as e:
        error_msg = str(e).lower()
        if 'invalid padding' in error_msg or 'unpad' in error_msg:
            return jsonify({
                "error": f"Gagal decrypt: Padding tidak valid. Kemungkinan:\n"
                         f"1. Key tidak sama dengan saat encrypt\n"
                         f"2. IV tidak sama dengan saat encrypt\n"
                         f"3. Ciphertext corrupt/dimodifikasi\n"
                         f"Detail: {str(e)}"
            }), 400
        else:
            return jsonify({"error": f"Gagal decrypt: {str(e)}"}), 400
    except Exception as e:
        print(f'Unexpected error in decrypt: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error: {str(e)}"}), 500


@app.route('/api/aes/image/encrypt', methods=['POST'])
def api_aes_encrypt_image():
    if 'image' not in request.files:
        return jsonify({"error": "Field image wajib ada"}), 400
    key_hex = request.form.get('key')
    iv_hex = request.form.get('iv')
    custom_sbox_json = request.form.get('sbox')  # Optional custom S-box
    
    if not key_hex:
        return jsonify({"error": "Key (hex) wajib diisi"}), 400

    img_file = request.files['image']
    if img_file.filename == '':
        return jsonify({"error": "Nama file kosong"}), 400

    filename = secure_filename(img_file.filename)
    plain_path = os.path.join(UPLOAD_FOLDER, filename)
    img_file.save(plain_path)

    try:
        cipher_path, iv_used = encrypt_file(plain_path, key_hex, AES_CIPHER_FOLDER, iv_hex)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    # If the uploaded file is an image, compute histogram/entropy and visualize ciphertext
    ext = os.path.splitext(filename)[1].lower().strip('.')
    is_img = ext in ALLOWED_IMAGES
    response = {
        "filename": os.path.basename(cipher_path),
        "iv": iv_used,
        "download_url": url_for('aes_cipher_file', filename=os.path.basename(cipher_path))
    }
    if is_img:
        try:
            # Plain image entropy & histogram
            response["entropy_plain"] = round(image_entropy(plain_path), 6)
            response["hist_plain"] = histogram_counts(plain_path)
            response["hist_rgb_plain"] = histogram_rgb(plain_path)

            # Build visualization from ciphertext bytes (grayscale)
            from PIL import Image
            import numpy as np
            im = Image.open(plain_path).convert('L')
            w, h = im.size
            with open(cipher_path, 'rb') as f:
                cb = f.read()
            
            # Repeat ciphertext if needed to fill entire image
            arr = np.frombuffer(cb, dtype=np.uint8)
            total_pixels = h * w
            if len(arr) < total_pixels:
                # Repeat ciphertext to fill the image
                repeats = (total_pixels // len(arr)) + 1
                arr = np.tile(arr, repeats)[:total_pixels]
            else:
                arr = arr[:total_pixels]
            
            vis = np.zeros((h, w), dtype=np.uint8)
            vis.flat[:] = arr
            vis_img_path = os.path.join(AES_VISUAL_FOLDER, f"vis_{os.path.splitext(filename)[0]}.png")
            Image.fromarray(vis, mode='L').save(vis_img_path)

            response["visual_url"] = url_for('aes_visual_file', filename=os.path.basename(vis_img_path))
            response["entropy_cipher"] = round(image_entropy(vis_img_path), 6)
            response["hist_cipher"] = histogram_counts(vis_img_path)

            # Build RGB visualization from ciphertext
            im_rgb = Image.open(plain_path).convert('RGB')
            w_rgb, h_rgb = im_rgb.size
            total_pixels_rgb = w_rgb * h_rgb * 3
            
            # Repeat ciphertext for RGB if needed
            arr_rgb = np.frombuffer(cb, dtype=np.uint8)
            if len(arr_rgb) < total_pixels_rgb:
                repeats = (total_pixels_rgb // len(arr_rgb)) + 1
                arr_rgb = np.tile(arr_rgb, repeats)[:total_pixels_rgb]
            else:
                arr_rgb = arr_rgb[:total_pixels_rgb]
            
            vis_rgb = np.zeros((h_rgb, w_rgb, 3), dtype=np.uint8)
            vis_rgb.flat[:] = arr_rgb
            vis_rgb_path = os.path.join(AES_VISUAL_FOLDER, f"vis_rgb_{os.path.splitext(filename)[0]}.png")
            Image.fromarray(vis_rgb, mode='RGB').save(vis_rgb_path)
            response["visual_rgb_url"] = url_for('aes_visual_file', filename=os.path.basename(vis_rgb_path))
            response["hist_rgb_cipher"] = histogram_rgb(vis_rgb_path)
            
            # If custom S-box provided, create S-box based visualization for comparison
            if custom_sbox_json:
                try:
                    custom_sbox = json.loads(custom_sbox_json)
                    sbox_cipher_name = f"sbox_cipher_{os.path.splitext(filename)[0]}.png"
                    sbox_cipher_path = os.path.join(ENCRYPTED_FOLDER, sbox_cipher_name)
                    apply_subbytes_to_image(plain_path, custom_sbox, sbox_cipher_path)
                    
                    response["sbox_visual_url"] = url_for('encrypted_file', filename=sbox_cipher_name)
                    response["sbox_entropy"] = round(image_entropy(sbox_cipher_path), 6)
                    response["sbox_hist"] = histogram_counts(sbox_cipher_path)
                    response["sbox_hist_rgb"] = histogram_rgb(sbox_cipher_path)
                    response["npcr_sbox"] = round(npcr(plain_path, sbox_cipher_path), 6)
                except Exception as e:
                    response["sbox_error"] = f"Custom S-box error: {str(e)}"

            # NPCR: modify a single pixel in plaintext, encrypt, visualize, compare
            mod_im = im.copy()
            px = mod_im.load()
            px[0,0] = (int(px[0,0]) + 1) % 256
            mod_plain_path = os.path.join(AES_PLAIN_FOLDER, f"mod_{filename}")
            mod_im.save(mod_plain_path)
            mod_cipher_path, _ = encrypt_file(mod_plain_path, key_hex, AES_CIPHER_FOLDER, iv_used)
            with open(mod_cipher_path, 'rb') as f:
                cb2 = f.read()
            arr2 = np.frombuffer(cb2, dtype=np.uint8)
            vis2 = np.zeros((h, w), dtype=np.uint8)
            n2 = min(h*w, arr2.size)
            vis2.flat[:n2] = arr2[:n2]
            vis_mod_img_path = os.path.join(AES_VISUAL_FOLDER, f"vis_mod_{os.path.splitext(filename)[0]}.png")
            Image.fromarray(vis2, mode='L').save(vis_mod_img_path)
            response["npcr"] = round(npcr(vis_img_path, vis_mod_img_path), 6)
            response["uaci"] = round(uaci(vis_img_path, vis_mod_img_path), 6)
        except Exception as e:
            response["analysis_error"] = str(e)

    return jsonify(response)


@app.route('/api/aes/image/decrypt', methods=['POST'])
def api_aes_decrypt_image():
    if 'cipher' not in request.files:
        return jsonify({"error": "Field cipher wajib ada"}), 400
    key_hex = request.form.get('key')
    iv_hex = request.form.get('iv')
    if not key_hex or not iv_hex:
        return jsonify({"error": "Key dan IV wajib diisi"}), 400
    original_ext = request.form.get('ext')

    cipher_file = request.files['cipher']
    if cipher_file.filename == '':
        return jsonify({"error": "Nama file kosong"}), 400

    filename = secure_filename(cipher_file.filename)
    cipher_path = os.path.join(AES_CIPHER_FOLDER, filename)
    cipher_file.save(cipher_path)

    # Check file size
    file_size = os.path.getsize(cipher_path)
    if file_size % 16 != 0:
        os.remove(cipher_path)
        return jsonify({"error": f"Cipher file size ({file_size} bytes) is not a multiple of 16. File may be corrupted. Make sure you're decrypting the correct .aes file with correct Key & IV."}), 400

    try:
        plain_path = decrypt_file(cipher_path, key_hex, AES_PLAIN_FOLDER, iv_hex, original_ext)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Decrypt error: {str(e)}"}), 400

    download_url = url_for('aes_plain_file', filename=os.path.basename(plain_path))
    return jsonify({
        "filename": os.path.basename(plain_path),
        "download_url": download_url
    })


@app.route('/api/explore-matrices', methods=['GET'])
def api_explore_matrices():
    """Explore multiple affine matrix candidates dan return ranked results"""
    n_candidates = request.args.get('n', default=50, type=int)
    top_n = request.args.get('top', default=10, type=int)
    seed = request.args.get('seed', default=None, type=int)
    
    # Limit untuk prevent overload
    if n_candidates > 200:
        n_candidates = 200
    if top_n > 50:
        top_n = 50
    
    results = explore_affine_candidates(n_candidates=n_candidates, seed=seed)
    top_results = get_top_candidates(results, top_n=top_n)
    
    # Simplify output (remove full sbox array for brevity)
    simplified = []
    for r in top_results:
        simplified.append({
            'id': r['id'],
            'bijective': r['bijective'],
            'balanced': r['balanced'],
            'sac': r['sac'],
            'diff_uniformity': r['diff_uniformity'],
            'nonlinearity': r['nonlinearity'],
            'matrix': r['matrix'],
            'constant': r['constant']
        })
    
    return jsonify({
        'total_tested': n_candidates,
        'top_n': top_n,
        'candidates': simplified
    })


@app.route('/api/download-candidate/<int:candidate_id>', methods=['POST'])
def api_download_candidate(candidate_id):
    """Generate & download specific candidate S-Box as Excel"""
    data = request.get_json()
    matrix = data.get('matrix')
    constant = data.get('constant')
    
    if not matrix or not constant:
        return jsonify({"error": "Matrix and constant required"}), 400
    
    import numpy as np
    from core.matrix_explorer import generate_sbox_from_affine
    
    M = np.array(matrix, dtype=int)
    C = np.array(constant, dtype=int)
    sbox = generate_sbox_from_affine(M, C)
    
    fname = f"candidate_{candidate_id}.xlsx"
    path = export_sbox_to_excel(sbox, filename=fname)
    
    return send_file(path, as_attachment=True, download_name=fname)


if __name__ == "__main__":
    app.run(debug=True, port=5000)

